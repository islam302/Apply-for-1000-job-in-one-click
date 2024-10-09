# -*- coding: utf-8 -*-
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from django.http import Http404, HttpResponse
from ChromeDriver import WebDriver
from selenium import webdriver
from bs4 import BeautifulSoup
from random import randint
import pandas as pd
import requests
import random
import json
import time
import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class LMSBot:

    def __init__(self):
        self.url = "https://bblms.kfu.edu.sa/"
        self.driver = None

    def start_driver(self):
        self.driver = WebDriver.start_driver(self)
        return self.driver

    def login(self, username, password):
        try:
            self.driver.get(self.url)

            username_field = self.driver.find_element(By.XPATH, '//*[@id="user_id"]')
            username_field.send_keys(username)

            password_field = self.driver.find_element(By.XPATH, '//*[@id="password"]')
            password_field.send_keys(password)

            wait = WebDriverWait(self.driver, 10)
            login_button = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'تسجيل الدخول')]")))

            login_button.click()
            time.sleep(1)
            login_button.click()

            accept1 = self.driver.find_element(By.XPATH, '//*[@id="btnYes"]')
            accept1.click()
            time.sleep(3)
        except :
            pass

    def get_subjects(self) -> list:
        div = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '.portletList-img.courseListing.coursefakeclass')))
        links = div.find_elements(By.TAG_NAME, "a")
        subjects = []
        for link in links:
            subject_name = link.text
            subject_url = link.get_attribute("href")
            # if subject_name  != "BBL-0827-111-CRN71408-Term144610-M: التفاضل والتكامل- طلاب":
            #     continue
            subjects.append({"name": subject_name, "url": subject_url})
        return subjects

    def classes_links(self) -> list:
        time.sleep(2)
        link_text = 'المحاضرات المسجلة والمحتوى الرقمي'
        link_element = self.driver.find_element(By.LINK_TEXT, link_text)
        link_element.click()

        time.sleep(3)
        ul = self.driver.find_element(By.CLASS_NAME, 'contentList')
        li = ul.find_elements(By.TAG_NAME, 'li')
        links = []
        del li[0:3]  # Removing the first three elements as you already do

        for element in li:
            try:
                div = element.find_element(By.CSS_SELECTOR, '.item.clearfix')
                link = div.find_element(By.TAG_NAME, 'h3').find_element(By.TAG_NAME, 'a').get_attribute('href')
                lecture_name = div.find_element(By.TAG_NAME, 'h3').text  # Extracting the lecture name

                # Check if the lecture name contains the word "الفصل"
                if "الفصل" in lecture_name:
                    links.append((link, lecture_name))  # Store as tuple (link, name)
            except Exception as e:
                continue
        return links

    def process_all_units(self, lecture_link):
        all_processed_links = []
        unit_links = self.get_units_links(lecture_link)
        for unit_link, unit_name in unit_links:  # Unpack unit links and names
            processed_unit_links = self.check_if_unit_has_parts(unit_link, unit_name)
            all_processed_links.extend(processed_unit_links)
        return all_processed_links

    def check_if_unit_has_parts(self, unit_link, unit_name):
        try:
            self.driver.get(unit_link)

            # Check if the unit has parts by looking for the content list container
            ul_parts = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="content_listContainer"]'))
            )

            # If found, extract all 'li' elements (parts)
            li_parts = ul_parts.find_elements(By.TAG_NAME, 'li')
            part_links = []

            for part in li_parts:
                try:
                    # Extract the label of the part (e.g., الفصل الاول - الوحدة الأولى - الجزء الأول)
                    part_label = part.find_element(By.TAG_NAME, 'h3').find_element(By.TAG_NAME, 'a').text
                    if 'الجزء' in part_label:
                        # Get the part link (href)
                        part_link = part.find_element(By.TAG_NAME, 'a').get_attribute('href')
                        part_links.append((part_link, part_label))  # Store as tuple (link, part name)
                except NoSuchElementException:
                    continue

            # If parts exist, return their links; if no parts, return the original unit link and its name
            if part_links:
                return part_links
            else:
                # Return original unit link with its name
                return [(unit_link, unit_name)]  # Return the unit link and its original name

        except TimeoutException:
            print(f"Timed out waiting for unit page to load: {unit_link}")
            return [(unit_link, unit_name)]  # Return with the original name on timeout

    def get_units_links(self, lecture_page_link):
        try:
            self.driver.get(lecture_page_link)
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.item.clearfix'))
            )
            elements = self.driver.find_elements(By.CSS_SELECTOR, '.item.clearfix')
            unit_links = []

            for element in elements:
                try:
                    # Check the unit name (text) before getting the link
                    unit_name = element.text

                    if "الوحدة" in unit_name:
                        # Get the unit link if it contains "الوحدة"
                        unit_link = element.find_element(By.TAG_NAME, 'a').get_attribute('href')
                        unit_links.append((unit_link, unit_name))  # Store as tuple (link, name)

                except (NoSuchElementException, StaleElementReferenceException) as e:
                    continue

            return unit_links

        except TimeoutException:
            return []

    def watch_video(self, link):
        try:
            try:
                self.check_secure_connection_error()
            except:
                pass
            self.driver.get(link)
            video_link = self.driver.find_element(By.ID, 'content_listContainer').find_elements(By.TAG_NAME, 'li')[3] \
                .find_element(By.TAG_NAME, 'a').get_attribute('href')
            self.driver.get(video_link)
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.TAG_NAME, 'video')))
            play_button = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.TAG_NAME, 'video')))

            if play_button:
                play_button.click()
                time.sleep(random.randint(6, 7))
                return True
            else:
                return False
        except:
            return False

    def listen_audio(self, link):
        try:
            try:
                self.check_secure_connection_error()
            except:
                pass
            self.driver.get(link)
            third_item_link = self.driver.find_element(By.ID, 'content_listContainer') \
                .find_elements(By.TAG_NAME, 'li')[2] \
                .find_element(By.TAG_NAME, 'a').get_attribute('href')
            self.driver.get(third_item_link)
            play_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "video"))
            )
            if play_button:
                time.sleep(random.randint(6, 7))
                return True
        except Exception as e:
            return False

    def logout(self):
        self.driver.get('https://bblms.kfu.edu.sa/')
        self.driver.find_element(By.ID, "topframe.logout.label").click()

    def login_report(self, username):
        try:
            file_exists = os.path.isfile("Login_Report.csv")
            if file_exists:
                df = pd.read_csv("Login_Report.csv", encoding="utf-8-sig")

                if username in df['اسم المستخدم'].values:
                    return
                else:
                    new_row = pd.DataFrame({"اسم المستخدم": [username], "بيانات تسجيل الدخول خاطئة": [""]})
                    df = pd.concat([pd.DataFrame(new_row), df], ignore_index=True, sort=False)
            else:
                df = pd.DataFrame(columns=["اسم المستخدم", "بيانات تسجيل الدخول خاطئة"])
                new_row = pd.DataFrame({"اسم المستخدم": [username], "بيانات تسجيل الدخول خاطئة": [""]})
                df = pd.concat([pd.DataFrame(new_row), df], ignore_index=True, sort=False)

            df.to_csv("Login_Report.csv", index=False, encoding="utf-8-sig")
        except Exception as e:
            pass

    def report(self, username, inter):
        try:
            file_path = "REPORT.xlsx"
            if os.path.isfile(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["اسم المستخدم", "الحالة"])

            username_exists = False
            for row in ws.iter_rows(values_only=True):
                if row[0] == username:
                    # Remove the prefix from the status
                    inter = [s.replace("BBL-0827-111-CRN71408-Term144610-M: ", "") for s in inter]
                    row = [username] + inter
                    username_exists = True
                    break

            if not username_exists:
                if inter:
                    # Remove the prefix from the status for new entries
                    inter = [s.replace("BBL-0827-111-CRN71408-Term144610-M: ", "") for s in inter]
                    row = [username] + inter
                else:
                    row = [username, "هذا الحساب جيد"]
                ws.append(row)

            # Set the column widths to accommodate long text
            for i in range(1, ws.max_column + 1):
                column = get_column_letter(i)
                ws.column_dimensions[column].width = 50  # Set to 50 or adjust as needed

            wb.save(file_path)

        except Exception as e:
            print(f"Error occurred: {e}")
            return False

    def check_secure_connection_error(self):
        error_message_element = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located(
                (By.XPATH, "//h1[contains(text(),'This site can't provide a secure connection')]"))
        )
        if error_message_element:
            logging.error('There is an error here. Please try again')
            time.sleep(3)
            self.driver.quit()
            sys.exit(1)

    def check_server_error(self):
        error_message_element = self.driver.find_element(By.XPATH,"//h2[contains(text(),'404 - File or directory not found.')]")
        if error_message_element:
            return True
        else:
            return False

    @staticmethod
    def data():
        try:
            data = pd.read_excel('users.xlsx')
            collage_numbers = data['الرقم الجامعي'].tolist()
            paswordat = data['الرقم السري'].tolist()
            usernames = collage_numbers
            passwords = paswordat
            login_data = zip(usernames, passwords)
            return dict(login_data)
        except:
            return False

    def main(self, number=0):
        users = self.data()
        if not users:
            print("Check your users.xlsx file")
            return False
        self.start_driver()

        for username, password in users.items():
            list_error = []
            self.login(username, password)
            try:
                if self.driver.current_url == "https://bblms.kfu.edu.sa/webapps/login/":
                    print(f"Credentials for {username} are incorrect. Skipping...")
                    self.login_report(username)
                    time.sleep(3)
                    continue
            except TimeoutException:
                pass
            try:
                accept_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="agree_button"]'))
                )
                accept_button.click()
            except TimeoutException:
                pass

            subjects = self.get_subjects()

            for subject in subjects:
                subject_name = subject["name"]
                subject_href = subject["url"]
                self.driver.get(subject_href)

                for link, lecture_name in self.classes_links()[-number:]:  # Unpack the link and lecture name
                    lecture_links = self.process_all_units(link)  # Now returns tuples (link, name)
                    self.driver.refresh()

                    for inner_lecture_link, inner_lecture_name in lecture_links:
                        audio_available = self.listen_audio(inner_lecture_link)
                        if not audio_available:
                            video_available = self.watch_video(inner_lecture_link)

                            if video_available:
                                list_error.append(
                                    f"{subject_name} >> {inner_lecture_name} لا يعمل (الملف الصوتي) ولكن يعمل الفيديو"
                                )
                                print(f"{subject_name} >> {inner_lecture_name} لا يعمل (الملف الصوتي) ولكن يعمل الفيديو")
                            else:
                                error = self.check_server_error()
                                if error:
                                    list_error.append(
                                        f"{subject_name} >> {inner_lecture_name} لا يعمل (الملف الصوتي والفيديو)"
                                    )
                                    print(
                                        f"{subject_name} >> {inner_lecture_name} لا يعمل (الملف الصوتي) ولكن يعمل الفيديو")

                                    continue
                                else:
                                    print(f'There is a problem with this user: {username}')
                                    list_error.append(
                                        f"{subject_name} >> {inner_lecture_name} لا يعمل (الملف الصوتي والفيديو)"
                                    )
                                    self.report(username, list_error)
                                    self.driver.quit()
                                    time.sleep(5)
                                    exit()

            self.logout()
            self.report(username, list_error)
            print(f"User {username} done...")
            time.sleep(5)

        print('All users done, have a nice day')
        time.sleep(2)
        self.driver.quit()


if __name__ == '__main__':
    bot = LMSBot()
    use = input("1- Start all Lectures ?\n2- Specific Classes ?\n")
    if use == "1":
        try:
            bot.main()
        except:
            print("There is an Error please try again")
    elif use == "2":
        x = int(input('Enter The Number Of Classes : \n'))
        try:
            bot.main(x)
        except:
            print("There is an Error please try again")
    else:
        os.system('exit')

