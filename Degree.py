# -*- coding: utf-8 -*-
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from openpyxl.styles import Alignment
from ChromeDriver import WebDriver
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from bs4 import BeautifulSoup
from random import randint
import pandas as pd
import requests
import random
import json
import time
import sys
import os
import uuid

class LMSBot:

    def __init__(self):
        self.url = "https://bblms.kfu.edu.sa/"
        self.driver = None

    def start_driver(self):
        self.driver = WebDriver.start_driver(self)
        return self.driver

    @staticmethod
    def data():
        try:
            data = pd.read_excel('users1.xlsx')
            collage_numbers = data['الرقم الجامعي'].tolist()
            paswordat = data['الرقم السري'].tolist()
            usernames = collage_numbers
            passwords = paswordat
            login_data = zip(usernames, passwords)
            return dict(login_data)
        except:
            return False

    def login(self, username, password):
        try:
            self.driver.get(self.url)

            username_field = self.driver.find_element(By.XPATH, '//*[@id="user_id"]')
            username_field.send_keys(username)

            password_field = self.driver.find_element(By.XPATH, '//*[@id="password"]')
            password_field.send_keys(password)

            login_button = self.driver.find_element(By.XPATH, "//button[contains(text(), 'تسجيل الدخول')]")
            login_button.click()

            accept1 = self.driver.find_element(By.XPATH, '//*[@id="btnYes"]')
            accept1.click()
        except :
            pass

    def logout(self):
        self.driver.get('https://bblms.kfu.edu.sa/')
        self.driver.find_element(By.ID, "topframe.logout.label").click()

    def login_report(self, username):
        try:
            file_exists = os.path.isfile("Login_Report.csv")
            if file_exists:
                df = pd.read_csv("Login_Report.csv", encoding="utf-8-sig")

                if username in df['اسم المستخدم'].values:
                    return
                else:
                    new_row = pd.DataFrame({"اسم المستخدم": [username], "بيانات تسجيل الدخول خاطئة": [""]})
                    df = pd.concat([pd.DataFrame(new_row), df], ignore_index=True, sort=False)
            else:
                df = pd.DataFrame(columns=["اسم المستخدم", "بيانات تسجيل الدخول خاطئة"])
                new_row = pd.DataFrame({"اسم المستخدم": [username], "بيانات تسجيل الدخول خاطئة": [""]})
                df = pd.concat([pd.DataFrame(new_row), df], ignore_index=True, sort=False)

            df.to_csv("Login_Report.csv", index=False, encoding="utf-8-sig")
        except Exception as e:
            print(f"حدث خطأ: {str(e)}")

    def get_subjects(self) -> list:
        div = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '.portletList-img.courseListing.coursefakeclass')))
        links = div.find_elements(By.TAG_NAME, "a")
        subjects = []
        for link in links:
            subject_name = link.text
            subject_url = link.get_attribute("href")
            subjects.append({"name": subject_name, "url": subject_url})
        return subjects

    def get_assignments(self, homework) -> list:
        choices = {
            'الواجب الأول': '1',
            'الواجب الاول': '1',
            'الواجب الثاني': '2',
            'الواجب الثالث': '3',
        }
        try:
            ul_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.ID, 'content_listContainer')))
            li_tags = ul_element.find_elements(By.TAG_NAME, 'li')
            assignments = []
            for li_tag in li_tags:
                assignment_name_element = li_tag.find_element(By.TAG_NAME, 'a')
                assignment_name = assignment_name_element.text.strip()
                if assignment_name in choices.keys() and choices[assignment_name] == homework:
                    assignment_url = assignment_name_element.get_attribute('href')
                    assignments.append({"name": assignment_name, "url": assignment_url})
                    break
            return assignments
        except:
            return False

    def take_screenshot(self, username):
        folder_path = os.path.join(os.getcwd(), str(username))
        os.makedirs(folder_path, exist_ok=True)
        screenshot_name = str(uuid.uuid4()) + ".png"
        screenshot_path = os.path.join(folder_path, screenshot_name)
        self.driver.save_screenshot(screenshot_path)

    def check_degree(self, username):
        try:
            name = "//a[contains(text(), 'View All Attempts')] | //a[contains(text(),' عرض كافة المحاولات')]"
            if name :
                link = self.driver.find_element(By.XPATH, name)
                link.click()

                table = WebDriverWait(self.driver, 20).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "attachments"))
                )
                rows = table.find_elements(By.TAG_NAME, "tr")

                last_grade = None

                for row in rows:
                    try:
                        link = row.find_element(By.TAG_NAME, "a")
                        if "review.jsp" in link.get_attribute("href"):
                            cells = row.find_elements(By.TAG_NAME, "td")
                            assignment_name = cells[1].text
                            if assignment_name == "الواجب الثالث":
                                grade_value = float(cells[-1].text)
                                if grade_value < 4:
                                    self.take_screenshot(username)
                                    return grade_value
                                return grade_value
                            else:
                                grade_value = float(cells[-1].text)
                                if last_grade is None or grade_value >= last_grade:
                                    last_grade = grade_value
                    except NoSuchElementException:
                        pass

                if last_grade is not None:
                    if last_grade >= 3:
                        return last_grade
                    else:
                        self.take_screenshot(username)
                        return last_grade
                else:
                    return
            else:
                return "هذا الواجل لم يحل"
        except:
            return "هذا الواجل لم يحل"

    def create_report(self, subjects, username, grades):
        file_path = "REPORT.xlsx"
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            headers = ["الرقم الجامعي"] + [f"مادة {i + 1}" for i in range(10)]
            ws.append(headers)

        row_data = [username]
        for subject in subjects:
            if subject["name"] in grades:
                row_data.append(grades[subject["name"]])
            else:
                row_data.append("لم يتم تحديد الدرجة أو الواجب لم يتم حله")
        ws.append(row_data)

        wb.save(file_path)

    def main(self):
        users = self.data()
        if not users:
            print("Check your users.xlsx file")
            return False

        choices = {
            'الواجب الأول': '1',
            'الواجب الاول': '1',
            'الواجب الثاني': '2',
            'الواجب الثالث': '3',
        }
        homework = input("homework [1] or [2] or [3]: ")

        self.start_driver()
        for username, password in users.items():
            grades = {}
            self.login(username, password)
            try:
                if self.driver.current_url == "https://bblms.kfu.edu.sa/webapps/login/":
                    print(f"Credentials for {username} are incorrect. Skipping...")
                    self.login_report(username)
                    continue
            except TimeoutException:
                pass
            try:
                accept_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="agree_button"]')))
                accept_button.click()
            except TimeoutException:
                pass
            # try:
            subjects = self.get_subjects()
            for subject in subjects:
                subject_name = subject["name"]
                subject_href = subject["url"]
                self.driver.get(subject_href)

                link_text = 'الواجبات والتمارين'
                link_element = self.driver.find_element(By.LINK_TEXT, link_text)
                link_element.click()
                assignments = self.get_assignments(homework)
                if not assignments:
                    print("No assignments found.")
                    continue

                for assignment in assignments:
                    assignment_name = assignment["name"]
                    assignment_href = assignment["url"]
                    self.driver.get(assignment_href)
                    begain = self.driver.find_element(By.CLASS_NAME, "submit")
                    begain.click()
                    grade = self.check_degree(username)
                    grades[subject_name] = grade

            self.create_report(subjects, username, grades)
            self.logout()
            print(f"User {username}  done...")
            time.sleep(5)
        print('All users Done, Have a nice day')
        time.sleep(2)
        self.driver.quit()

if __name__ == '__main__':
    bot = LMSBot()
    try:
        bot.main()
    except:
        print("There is something wrong with please check if our files opened or internet cutten and try again")
        time.sleep(10)

