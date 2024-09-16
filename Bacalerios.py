from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from ChromeDriver import WebDriver
import pandas as pd
import openpyxl
import requests
import time
import sys
import os



class LMSBot:

    def __init__(self):
        self.url = "https://del-vls.kfu.edu.sa/Login.aspx"
        self.driver = None

    def start_driver(self):
        self.driver = WebDriver.start_driver(self)
        return self.driver

    @staticmethod
    def data():
        try:
            data = pd.read_excel('f.xlsx')
            collage_numbers = data['الرقم الجامعي'].tolist()
            passwords = data['الرقم السري'].tolist()
            login_data = dict(zip(collage_numbers, passwords))
            return login_data

        except Exception as e:
            print(f"Error reading users.xlsx: {str(e)}")
            return False

    def login(self, username, password):
        try:
            self.driver.get(self.url)

            username_field = self.driver.find_element(By.XPATH, "//input[@name='txt_userID']")
            username_field.send_keys(username)

            password_field = self.driver.find_element(By.XPATH, "//input[@name='txt_password']")
            password_field.send_keys(password)

            login_button = self.driver.find_element(By.XPATH, "//button[@id='btn_Login']")
            login_button.click()
        except Exception as e:
            print(f"Error occurred while logging in: {str(e)}")

    def login_report(self, username):
        try:
            file_exists = os.path.isfile("Login_Report.csv")
            if file_exists:
                df = pd.read_csv("Login_Report.csv", encoding="utf-8-sig")

                if username in df['اسم المستخدم'].values:
                    return
                else:
                    new_row = pd.DataFrame({"اسم المستخدم": [username], "بيانات تسجيل الدخول خاطئة": [""]})
                    df = pd.concat([pd.DataFrame(new_row), df], ignore_index=True, sort=False)
            else:
                df = pd.DataFrame(columns=["اسم المستخدم", "بيانات تسجيل الدخول خاطئة"])
                new_row = pd.DataFrame({"اسم المستخدم": [username], "بيانات تسجيل الدخول خاطئة": [""]})
                df = pd.concat([pd.DataFrame(new_row), df], ignore_index=True, sort=False)

            df.to_csv("Login_Report.csv", index=False, encoding="utf-8-sig")
        except Exception as e:
            pass

    def check_fees(self):
        if "https://del-vls.kfu.edu.sa/fees.aspx" in self.driver.current_url:
            return True
        else:
            return False

    def get_subjects(self) -> list:
        subjects = []
        try:
            div = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.ID, "MainBody_Panel1")))
            links = div.find_elements(By.TAG_NAME, "a")
            for link in links:
                try:
                    subject_name = link.text
                    subject_url = link.get_attribute("href")
                    subjects.append({"name": subject_name, "url": subject_url})
                except Exception as e:
                    print(f"Error processing subject {subject_name}: {e}")
                    continue
        except Exception as e:
            print(f"Error getting subjects: {e}")
        return subjects

    def get_lectures(self) -> list:
        container = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, "MainBody_Lect_Panel"))
        )
        links = container.find_elements(By.TAG_NAME, "a")
        lecture_links = []
        for link in links:
            name = link.text
            link = link.get_attribute("href")
            lecture_links.append((name, link))
        return lecture_links

    def audio(self):
        try:
            element = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.LINK_TEXT, "تحميل المحاضرة بصيغة الصوت"))
            )
            element.click()

            time.sleep(7)
            self.close_current_tab()
            return True
        except:
            return False

    def video(self):
        try:
            time.sleep(2)
            element = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.LINK_TEXT, "تحميل المحاضرة بصيغة الفيديو"))
            )
            element.click()
            time.sleep(7)
            self.close_current_tab()
            return True
        except:
            return False

    def report(self, username, inter):
        try:
            file_path = "REPORT.xlsx"
            if os.path.isfile(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["اسم المستخدم", "الحالة"])

            username_exists = False
            for row in ws.iter_rows(values_only=True):
                if row[0] == username:
                    row = [username] + inter
                    username_exists = True
                    break

            if not username_exists:
                if inter:
                    row = [username] + inter
                else:
                    row = [username, "هذا الحساب جيد"]
                ws.append(row)
            wb.save(file_path)

        except Exception as e:
            print("Error:", e)

    def logout(self):
        self.driver.find_element(By.ID, "StdHeaderController1_LogOut_lbtn").click()

    def main(self, number=0):
        users = self.data()
        if not users:
            print("Check your users.xlsx file")
            return False

        for username, password in users.items():
            list_error = []
            self.start_driver()

            self.login(username, password)
            check = self.check_fees()
            if check:
                list_error.append(f"{username} has not paid the fees")
                self.report(username, list_error)
                print(f"User {username} has not paid the fees. Skipping...")
                self.driver.quit()
                time.sleep(5)
                continue

            try:
                if self.driver.current_url == "https://bblms.kfu.edu.sa/webapps/login/":
                    print(f"Credentials for {username} are incorrect. Skipping...")
                    self.driver.quit()
                    self.login_report(username)
                    continue
            except TimeoutException:
                pass

            try:
                accept_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="agree_button"]')))
                accept_button.click()
            except TimeoutException:
                pass

            subjects = self.get_subjects()
            for subject in subjects:
                subject_name = subject["name"]
                subject_href = subject["url"]
                self.driver.execute_script(subject_href)

                link_text = "المحاضرات المسجلة  والمحتوى"
                link = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//*[contains(text(), '{link_text}')]"))
                )
                link.click()

                lectures = self.get_lectures()[-number:]
                print(lectures)
                for lecture in lectures:
                        lecture_name = lecture[0]
                        lecture_link = lecture[1]

                        self.driver.execute_script(lecture_link)

                        audio_available = self.audio()

                        if not audio_available:
                            video_available = self.video()

                            if video_available:
                                list_error.append(
                                    f"{subject_name} >> {lecture_name} لا يعمل (الملف الصوتي) ولكن يعمل الفيديو")
                            else:
                                print(f'There is a Problem in this user : {username}')
                                list_error.append(
                                    f"{subject_name} >> {lecture_name} لا يعمل (الملف الصوتي والفيديو)")
                                self.driver.quit()
                                self.report(username, list_error)
                                time.sleep(5)
                                exit()

            self.logout()
            self.report(username, list_error)
            print(f"User {username} done...")
            self.driver.quit()
            time.sleep(5)

        print('All users Done, Have a nice day')
        time.sleep(2)
        self.driver.quit()


if __name__ == '__main__':
    bot = LMSBot()
    num = input('1-All Subjects\n2-Specific Subjects\n')
    if num == '1':
        bot.main()
    elif num == '2':
        n = int(input('Enter number of lectures : \n'))
        bot.main(n)
    else:
        sys.exit()



